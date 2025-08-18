# -*- coding: utf-8 -*-
"""
Created on Fri Jun 26 11:50:09 2020

@author: taein78.kim
"""
import time
import datetime as dt
import logging
import pandas as pd
import datetime as dt
import pymysql
from openpyxl import Workbook
from openpyxl import load_workbook
from urllib.request import urlopen
from bs4 import BeautifulSoup

import sqlalchemy
from sqlalchemy import create_engine
from sqlalchemy.orm import scoped_session, sessionmaker
from sqlalchemy.ext.declarative import declarative_base
#from ConnecttingValue import password, id, host, db

# logger ####################################################################################################
# create logger with module name
logger_dbms = logging.getLogger('main.dbms')

global g_table_columns_machine
global g_table_columns_node
global g_table_columns_station

db_name='mcc'
g_table_columns_machine={'abnormal_stop':['abnormal_stop_timeGT'],
                         'amc_comm':['amc_comm_failGT'],
                         'amc_voltage':['amc_3VGT','amc_3VLT','amc_5VGT','amc_5VLT',
                                        'amc_12VpGT','amc_12VpLT','amc_12VmGT','amc_12VmLT'],
                         'bcr_drive':['BCR_drive_readfail_countGT','BCR_drive_trigger_countGT'],
                         'bcr_trans':['BCR_trans_readfail_countGT','BCR_trans_trigger_countGT'],
                         'cid_link':['CID_link_distance_diffGT','CID_link_distance_diffLT'],
                         'cido_monitor':['CIDO_first_linklevelLT','CIDO_fiber_errorGT','CIDO_RF_errorGT'],
                         'cidr_monitor':['CIDR_first_linklevelLT','CIDR_fiber_errorGT','CIDR_RF_errorGT',
                                         'CIDR_max_currentGT','CIDR_max_voltageGT',
                                         'CIDR_max_temperatureGT','CIDR_noiseGT'],
                         'drive_follow':['drive_follow_errorGT'],
                         'drive_toque':['drive_toque_frontGT','drive_toque_frontLT',
                                        'drive_toque_rearGT','drive_toque_rearLT'],
                         'fan_fail':['fan_fail_countGT'],
                         'foup_cover_detect_position':['foup_cover_detect_on_positionGT'],
                         'hand_comm':['hand_comm_failGT'],
                         'hoist_follow':['hoist_follow_errorGT','hoist_follow_errorLT'],                         
                         'hoist_home':['hoist_home_off_position_ldGT','hoist_home_off_position_ldLT',
                                       'hoist_home_on_position_ldGT','hoist_home_on_position_ldLT',
                                       'hoist_home_off_position_uldGT','hoist_home_off_position_uldLT',
                                       'hoist_home_on_position_uldGT','hoist_home_on_position_uldLT'],
                         'inner_foup_detect':['foup_abnormal_detectGT','foup_abnormal_undetectedGT'],
                         'inner_foup_detect_positon':['foup_detect_off_positionGT','foup_detect_off_positionLT',
                                                      'foup_detect_on_positionGT','foup_detect_on_positionLT'],
                         'ipc_monitor':['ipc_temperatureGT','ipc_voltageGT','ipc_voltageLT'],
                         'ipc_performance':['ipc_cpu_usageGT','ipc_memory_usageGT'],
                         'lookdown':['lookdown_detect_timeGT'],
                         'mark_node':['node_mark_distanceGT','node_mark_distanceLT',
                                      'node_sensor_distanceGT','node_sensor_distanceLT',
                                      'node_mark_enter_velocityGT'],
                         'mark_qr':['QR_mark_distanceGT','QR_mark_distanceLT',
                                    'QR_offsetGT','QR_offsetLT',
                                    'QR_mark_enter_velocityGT',
                                    'QR_node_offset_diffGT','QR_node_offset_diffLT'],
                         'mark_station':['station_mark_distanceGT','station_mark_distanceLT',
                                         'station_sensor_distanceGT','station_sensor_distanceLT',
                                         'station_mark_enter_velocityGT','station_sensor_haunt_countGT',
                                         'station_node_offset_diffGT','station_node_offset_diffLT'],
                         'oscillation':['oscillation_countGT','oscillation_pause_timeGT'],
                         'regulator_current':['regulator_current_IoutGT','regulator_current_IoutLT',
                                              'regulator_current_IacGT','regulator_current_IacLT'],
                         'regulator_temperature':['regulator_temperature_pickupGT','regulator_temperature_IGBGT',
                                                  'regulator_temperature_EcapGT','regulator_temperature_Ecap2GT',
                                                  'regulator_temperature_EcapRoundGT',
                                                  'regulator_temperature_diff_EcapGT','regulator_temperature_diff_Ecap2GT'],
                         'regulator_voltage':['regulator_voltage_EdcGT','regulator_voltage_EdcLT',
                                              'regulator_voltage_EcapGT','regulator_voltage_EcapLT',
                                              'regulator_voltage_EoutGT','regulator_voltage_EoutLT',
                                              'regulator_voltage_diff_EdcEoutGT','regulator_voltage_diff_EdcEoutLT',
                                              'regulator_voltage_diff_EdcEcapGT','regulator_voltage_diff_EdcEcapLT'],
                         'rotate_sensor':['rotate_sensor_0_on_degreeGT','rotate_sensor_0_off_degreeGT',
                                          'rotate_sensor_180_on_degreeLT','rotate_sensor_180_off_degreeLT'],
                         'rssi':['rssi_levelGT'],
                         'shutter_time':['shutter_front_open_movetimeGT','shutter_front_open_movetimeLT',
                                         'shutter_front_close_movetimeGT','shutter_front_close_movetimeLT',
                                         'shutter_rear_open_movetimeGT','shutter_rear_open_movetimeLT',
                                         'shutter_rear_close_movetimeGT','shutter_rear_close_movetimeLT'],
                         'slide_follow':['slide_follow_errorGT','slide_follow_errorLT'],
                         'slide_home':['slide_home_off_position_lstbGT','slide_home_off_position_lstbLT',
                                       'slide_home_off_position_rstbGT','slide_home_off_position_rstbLT',
                                       'slide_home_on_position_lstbGT','slide_home_on_position_lstbLT',
                                       'slide_home_on_position_rstbGT','slide_home_on_position_rstbLT'],
                         'slide_home_off':['slide_home_sensor_offGT'],
                         'slide_position':['slide_uld_position_errorGT','slide_uld_position_errorLT',
                                           'slide_ld_position_errorGT','slide_ld_position_errorLT'],
                         'steer_sensor':['steer_front_left_sensor_countGT','steer_front_right_sensor_countGT',
                                         'steer_rear_left_sensor_countGT','steer_rear_right_sensor_countGT'],
                         'steer_time':['steer_front_left_movetimeGT','steer_front_left_movetimeLT',
                                       'steer_front_right_movetimeGT','steer_front_right_movetimeLT',
                                       'steer_rear_left_movetimeGT','steer_rear_left_movetimeLT',
                                       'steer_rear_right_movetimeGT','steer_rear_right_movetimeLT'],
                         'tag_node':['node_compensationGT','node_compensationLT'],
                         'thread_cycle':['thread_cycle_statusGT','thread_cycle_amcGT','thread_cycle_taskGT',
                                         'thread_cycle_excuteGT','thread_cycle_logGT','thread_cycle_extrajobGT'],
                         # add - 2020.09.18
                         'hand_movetime':['hand_release_movetimeGT','hand_release_movetimeLT',
                                          'hand_grip_movetimeGT','hand_grip_movetimeLT'],
                         'hoist_home_sensor':['hoist_home_sensor_off_countGT'],
                         'obs_abnormal_detect':['obs_abnormal_detect_timeGT']
                         }

g_table_columns_node={'amc_voltage':['amc_3VGT','amc_3VLT','amc_5VGT','amc_5VLT',
                                     'amc_12VpGT','amc_12VpLT','amc_12VmGT','amc_12VmLT'],
                      'bcr_drive':['BCR_drive_readfail_countGT','BCR_drive_trigger_countGT'],
                      'cid_link':['CID_link_distance_diffGT','CID_link_distance_diffLT'],
                      'cido_monitor':['CIDO_first_linklevelLT','CIDO_fiber_errorGT','CIDO_RF_errorGT'],
                      'cidr_monitor':['CIDR_first_linklevelLT','CIDR_fiber_errorGT','CIDR_RF_errorGT',
                                      'CIDR_max_currentGT','CIDR_max_voltageGT',
                                      'CIDR_max_temperatureGT','CIDR_noiseGT'],
                      'drive_follow':['drive_follow_errorGT'],
                      'drive_toque':['drive_toque_frontGT','drive_toque_frontLT',
                                     'drive_toque_rearGT','drive_toque_rearLT'],
                      'hand_comm':['hand_comm_failGT'],
                      'hoist_follow':['hoist_follow_errorGT','hoist_follow_errorLT'],
                      'mark_node':['node_mark_distanceGT','node_mark_distanceLT',
                                   'node_sensor_distanceGT','node_sensor_distanceLT',
                                   'node_mark_enter_velocityGT'],
                      'regulator_current':['regulator_current_IoutGT','regulator_current_IoutLT',
                                           'regulator_current_IacGT','regulator_current_IacLT'],
                      'regulator_temperature':['regulator_temperature_pickupGT','regulator_temperature_IGBGT',
                                               'regulator_temperature_EcapGT','regulator_temperature_Ecap2GT',
                                               'regulator_temperature_EcapRoundGT',
                                               'regulator_temperature_diff_EcapGT','regulator_temperature_diff_Ecap2GT'],
                      'regulator_voltage':['regulator_voltage_EdcGT','regulator_voltage_EdcLT',
                                           'regulator_voltage_EcapGT','regulator_voltage_EcapLT',
                                           'regulator_voltage_EoutGT','regulator_voltage_EoutLT',
                                           'regulator_voltage_diff_EdcEoutGT','regulator_voltage_diff_EdcEoutLT',
                                           'regulator_voltage_diff_EdcEcapGT','regulator_voltage_diff_EdcEcapLT'],
                      'rssi':['rssi_levelGT'],
                      'slide_follow':['slide_follow_errorGT','slide_follow_errorLT'],
                      'slide_home_off':['slide_home_sensor_offGT'],
                      'steer_sensor':['steer_front_left_sensor_countGT','steer_front_right_sensor_countGT',
                                      'steer_rear_left_sensor_countGT','steer_rear_right_sensor_countGT'],
                      'steer_time':['steer_front_left_movetimeGT','steer_front_left_movetimeLT',
                                    'steer_front_right_movetimeGT','steer_front_right_movetimeLT',
                                    'steer_rear_left_movetimeGT','steer_rear_left_movetimeLT',
                                    'steer_rear_right_movetimeGT','steer_rear_right_movetimeLT'],
                      'tag_node':['node_compensationGT','node_compensationLT'],
                      # add - 2020.09.18
                      'obs_abnormal_detect':['obs_abnormal_detect_timeGT'],
                      }

g_table_columns_station={'bcr_trans':['BCR_trans_readfail_countGT','BCR_trans_trigger_countGT'],
                         'lookdown':['lookdown_detect_timeGT'],
                         'mark_qr':['QR_mark_distanceGT','QR_mark_distanceLT',
                                    'QR_offsetGT','QR_offsetLT',
                                    'QR_mark_enter_velocityGT',
                                    'QR_node_offset_diffGT','QR_node_offset_diffLT'],
                         'mark_station':['station_mark_distanceGT','station_mark_distanceLT',
                                         'station_sensor_distanceGT','station_sensor_distanceLT',
                                         'station_mark_enter_velocityGT','station_sensor_haunt_countGT',
                                         'station_node_offset_diffGT','station_node_offset_diffLT'],
                         'oscillation':['oscillation_countGT','oscillation_pause_timeGT'],                         
                         # add - 2020.09.18
                         'hoist_home_sensor':['hoist_home_sensor_off_countGT']                      
                         }

## 0. Connect #################################################################
# Open database connection
def connect(db_name=""):
    if(len(db_name)>0):
        conn = pymysql.connect(host='localhost',
                             port=3306,
                             user='root',
                             passwd='semes_12',
                             db=db_name,
                             charset='utf8',
                             autocommit=True)
    else:
        conn = pymysql.connect(host='localhost',
                             port=3306,
                             user='root',
                             passwd='semes_12',
                             charset='utf8')
        
    return conn
     

## Etc #################################################################
# get db version
def get_version():
    # db connect
    conn = connect()
     
    # prepare a cursor object using cursor() method
    cursor = conn.cursor()
     
    # execute SQL query using execute() method.
    cursor.execute("SELECT VERSION()")
     
    # Fetch a single row using fetchone() method.
    data = cursor.fetchone()
     
    logger_dbms.info("Database version : %s ", data)
     
    # disconnect from server
    conn.close()


# get table column names set
def get_table_columns(db_name, table_name):
    # db connect
    conn = connect(db_name)  
    sql='SHOW COLUMNS from '+table_name
 
    df=pd.read_sql_query(sql, conn)
    cols=df['Field'].tolist()
    
    return(pd.DataFrame(columns=cols))

# get group names
def get_table_group(db_name, group):
    if(group == 'machine'):
        return g_table_columns_machine
    if(group == 'node'):
        return g_table_columns_node
    if(group == 'station'):
        return g_table_columns_station

# read table to dataframe
def read_table2df(db_name, table_name):
    # db connect
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
 
    df=pd.read_sql_table(table_name, con=engine)
    
    return df


def insert_table_column(db_name, table_name, col_name):
    # db connect
    conn = connect(db_name)
    logger_dbms.info(f'Add column: {table_name} -> {col_name}')
    try:
        with conn.cursor() as cursor:
            sql='ALTER TABLE ' + table_name + ' ADD COLUMN ' + col_name + ' FLOAT(12) NOT NULL DEFAULT 0'
            cursor.execute(sql)
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()    

## 1. _###################################################################

# create database
def create_db(db_name):
    # db connect
    conn = connect()
    
    try:
        with conn.cursor() as cursor:
            sql="CREATE DATABASE IF NOT EXISTS "+ db_name
            cursor.execute(sql)
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()

    
# create table    
def create_table_history(db_name):
    # db connect
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `history` (
                	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                	`site` varchar(32) NOT NULL DEFAULT '' COMMENT '위치 ' COLLATE 'utf8_general_ci',
                	`site_group` varchar(64) NULL DEFAULT '' COMMENT '위치 상세' COLLATE 'utf8_general_ci',
                	`term_from` CHAR(10) NOT NULL DEFAULT '' COMMENT '기간 시작' COLLATE 'utf8_general_ci',
                	`term_to` CHAR(10) NOT NULL DEFAULT '' COMMENT '기간 완료' COLLATE 'utf8_general_ci',
                    `amount_trans_eq` SMALLINT(5) UNSIGNED NOT NULL COMMENT 'EQ 이적재 회수',
                    `amount_trans_stk` SMALLINT(5) UNSIGNED NOT NULL COMMENT 'Stocker 이적재 회수',
                    `amount_trans_lstb` SMALLINT(5) UNSIGNED NOT NULL COMMENT 'LSTB 이적재 회수',
                    `amount_trans_rstb` SMALLINT(5) UNSIGNED NOT NULL COMMENT 'RSTB 이적재 회수',
                	`diag_machine_total` SMALLINT(5) UNSIGNED NOT NULL COMMENT '호기진단 항목 수',
                	`diag_machine_detect` SMALLINT(5) UNSIGNED NOT NULL COMMENT '호기진단 검출 수',
                    `diag_node_total` SMALLINT(5) UNSIGNED NOT NULL COMMENT '노드진단 항목 수',
                	`diag_node_detect` SMALLINT(5) UNSIGNED NOT NULL COMMENT '노드진단 검출 수',
                    `diag_station_total` SMALLINT(5) UNSIGNED NOT NULL COMMENT '포트진단 항목 수',
                	`diag_station_detect` SMALLINT(5) UNSIGNED NOT NULL COMMENT '포트진단 검출 수',
                	`comment` VARCHAR(1024) NULL DEFAULT NULL COMMENT '주석 ' COLLATE 'utf8_general_ci',                    
                	PRIMARY KEY (`analysis_time`) USING BTREE
                )
                COMMENT='MCC 진단 이력 테이블'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()
        
# create table    
def create_table_index_machine(db_name):
    # db connect
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `index_machine` (
                	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                	`id_machine` char(6) NOT NULL DEFAULT '' COMMENT 'OHT Name' COLLATE 'utf8_general_ci',
                	`total` FLOAT NOT NULL COMMENT '종합 건강도',
                    `abnormal_stop` FLOAT NOT NULL COMMENT '무언정지',
                    `amc_comm` FLOAT NOT NULL COMMENT 'amc 통신',
                    `amc_voltage` FLOAT NOT NULL COMMENT 'amc 전압',
                    `bcr_drive` FLOAT NOT NULL COMMENT '주행 Bar Code Read',
                    `bcr_trans` FLOAT NOT NULL COMMENT '이적재 Bar Code Read',
                    `cid_link` FLOAT NOT NULL COMMENT 'CID Link',
                    `cido_monitor` FLOAT NOT NULL COMMENT 'CID-O 모니터 데이터',
                    `cidr_monitor` FLOAT NOT NULL COMMENT 'CID-R 모니터 데이터',
                    `drive_follow` FLOAT NOT NULL COMMENT '주행 엔코더 오차',
                    `drive_toque` FLOAT NOT NULL COMMENT '주행 중 토크',
                    `fan_fail` FLOAT NOT NULL COMMENT 'Fan 상태',
                    `foup_cover_detect_position` FLOAT NOT NULL COMMENT 'Foup Cover 센서 조정 상태',
                    `hand_comm` FLOAT NOT NULL COMMENT 'Hand 통신',
                    `hoist_follow` FLOAT NOT NULL COMMENT 'Hoist 엔코더 오차',
                    `hoist_home` FLOAT NOT NULL COMMENT 'Hoist Home 조정 상태',
                    `inner_foup_detect` FLOAT NOT NULL COMMENT 'Foup 감지 상태',
                    `inner_foup_detect_positon` FLOAT NOT NULL COMMENT 'Foup 감지 센서 조정 상태',
                    `ipc_monitor` FLOAT NOT NULL COMMENT 'IPC 보드 상태',
                    `ipc_performance` FLOAT NOT NULL COMMENT 'IPC 성능',
                    `lookdown` FLOAT NOT NULL COMMENT 'Lookdown',
                    `mark_node` FLOAT NOT NULL COMMENT '주행 마크 분석',
                    `mark_qr` FLOAT NOT NULL COMMENT 'QR 마크 분석',
                    `mark_station` FLOAT NOT NULL COMMENT '이적재 마크 분석',
                    `oscillation` FLOAT NOT NULL COMMENT '오실레이션',
                    `regulator_current` FLOAT NOT NULL COMMENT '레귤레이터 전류',
                    `regulator_temperature` FLOAT NOT NULL COMMENT '레귤레이터 온도',
                    `regulator_voltage` FLOAT NOT NULL COMMENT '레귤레이터 전압',
                    `rotate_sensor` FLOAT NOT NULL COMMENT '로테이트 위치',
                    `rssi` FLOAT NOT NULL COMMENT '무선감도',
                    `shutter_time` FLOAT NOT NULL COMMENT '셔터 구동시간',
                    `slide_follow` FLOAT NOT NULL COMMENT '슬라이드 엔코더 오차',
                    `slide_home` FLOAT NOT NULL COMMENT '슬라이드 홈 조정 상태',
                    `slide_home_off` FLOAT NOT NULL COMMENT '주행 중 슬라이드 홈 벗어남',
                    `slide_position` FLOAT NOT NULL COMMENT '슬라이드 제어 오차',
                    `steer_sensor` FLOAT NOT NULL COMMENT '조향 센서',
                    `steer_time` FLOAT NOT NULL COMMENT '조향 시간',
                    `tag_node` FLOAT NOT NULL COMMENT '주행 노드 오차',
                    `thread_cycle` FLOAT NOT NULL COMMENT '제어 사이클',
                    
                    `hand_movetime` FLOAT NOT NULL COMMENT 'hand 구동 시간',
                    `hoist_home_sensor` FLOAT NOT NULL COMMENT 'hoist 원점 위치',
                    `obs_abnormal_detect` FLOAT NOT NULL COMMENT '장애물 오감지',
                         
                	INDEX `id_machine` (`id_machine`) USING BTREE,
                	INDEX `analysis_time` (`analysis_time`) USING BTREE
                )
                COMMENT='Machine 기준 진단 결과 테이블'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()


# create table    
def create_table_index_node(db_name):
# db connect
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `index_node` (
                	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                	`id_node` char(6) NOT NULL DEFAULT '' COMMENT 'Node Number' COLLATE 'utf8_general_ci',                	
                	`total` FLOAT NOT NULL COMMENT '종합 건강도',
                    `amc_voltage` FLOAT NOT NULL COMMENT 'amc 전압',
                    `bcr_drive` FLOAT NOT NULL COMMENT '주행 Bar Code Read',
                    `cid_link` FLOAT NOT NULL COMMENT 'CID Link',
                    `cido_monitor` FLOAT NOT NULL COMMENT 'CID-O 모니터 데이터',
                    `cidr_monitor` FLOAT NOT NULL COMMENT 'CID-R 모니터 데이터',
                    `drive_follow` FLOAT NOT NULL COMMENT '주행 엔코더 오차',
                    `drive_toque` FLOAT NOT NULL COMMENT '주행 중 토크',
                    `hand_comm` FLOAT NOT NULL COMMENT 'Hand 통신',
                    `hoist_follow` FLOAT NOT NULL COMMENT 'Hoist 엔코더 오차',
                    `mark_node` FLOAT NOT NULL COMMENT '주행 마크 분석',
                    `regulator_current` FLOAT NOT NULL COMMENT '레귤레이터 전류',
                    `regulator_temperature` FLOAT NOT NULL COMMENT '레귤레이터 온도',
                    `regulator_voltage` FLOAT NOT NULL COMMENT '레귤레이터 전압',
                    `rssi` FLOAT NOT NULL COMMENT '무선감도',
                    `slide_follow` FLOAT NOT NULL COMMENT '슬라이드 엔코더 오차',
                    `slide_home_off` FLOAT NOT NULL COMMENT '주행 중 슬라이드 홈 벗어남',
                    `steer_sensor` FLOAT NOT NULL COMMENT '조향 센서',
                    `steer_time` FLOAT NOT NULL COMMENT '조향 시간',
                    `tag_node` FLOAT NOT NULL COMMENT '주행 노드 오차',
                    
                    `obs_abnormal_detect` FLOAT NOT NULL COMMENT '장애물 오감지',
                   
                	INDEX `id_node` (`id_node`) USING BTREE,
                	INDEX `analysis_time` (`analysis_time`) USING BTREE
                )
                COMMENT='Node 기준 진단 결과 테이블'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()
        

# create table
def create_table_index_station(db_name):
# db connect
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `index_station` (
                	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                	`id_station` char(6) NOT NULL DEFAULT '' COMMENT 'Station No' COLLATE 'utf8_general_ci',                	
                	`total` FLOAT NOT NULL COMMENT '종합 건강도',
                    `bcr_trans` FLOAT NOT NULL COMMENT '이적재 Bar Code Read',
                    `lookdown` FLOAT NOT NULL COMMENT 'Lookdown',
                    `mark_qr` FLOAT NOT NULL COMMENT 'QR 마크 분석',
                    `mark_station` FLOAT NOT NULL COMMENT '이적재 마크 분석',
                    `oscillation` FLOAT NOT NULL COMMENT '오실레이션',
                    
                    `hoist_home_sensor` FLOAT NOT NULL COMMENT 'hoist 원점 위치',
                   
                	INDEX `id_station` (`id_station`) USING BTREE,
                	INDEX `analysis_time` (`analysis_time`) USING BTREE
                )
                COMMENT='Station 기준 진단 결과 테이블'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()
  

# create table
def create_table_ivalue_machine(db_name, table_name, values):
    if((len(table_name)>0) & (len(values)>0)):
        # db connect
        conn = connect(db_name)
        
        try:
            with conn.cursor() as cursor:
                sql='CREATE TABLE IF NOT EXISTS ivalue_machine_'+ table_name + ''' (
                    	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                    	`id_machine` char(6) NOT NULL DEFAULT '' COMMENT 'OHT Name' COLLATE 'utf8_general_ci',                	
                    	`total` FLOAT NOT NULL COMMENT '종합 건강도','''
                        
                for v in values:
                    sql += v + ' FLOAT NOT NULL,'
                        
                sql += ''' INDEX `id_machine` (`id_machine`) USING BTREE,
                    	INDEX `analysis_time` (`analysis_time`) USING BTREE
                        )
                        COMMENT='machine 기준 진단 그룹 상세 결과 테이블'
                        COLLATE='utf8_general_ci'
                        ENGINE=InnoDB
                       '''
                cursor.execute(sql)        
            conn.commit()
        except Exception as e:
            logger_dbms.exception(e)
        finally:        
            conn.close()
    else:
        logger_dbms.warning('invalid para. length: table_name(%d), values(%d)', len(table_name), len(values))
        logger_dbms.warning('invalid para. length: table_name(%d), values(%d)', len(table_name), len(values))



def set_ivalue_machine_table(db_name):
    table_names=g_table_columns_machine

    for k,v in table_names.items():
        create_table_ivalue_machine(db_name, k, v)

def update_machine_table(db_name):
    '''
    Notice: . g_table_columns_machine 에 추가할 항목을 우선 업데이트 후 실행 필요
            . 추가 업데이트만 가능
              g_# list에서 현재 db 항목 삭제 후 남은(추가 할) 항목 table 생성
    '''
    table_name='index_machine'
    col_lst=get_table_columns(db_name, table_name).columns.to_list()
    del(col_lst[0:3]) # analysis_time, id_machine, total
    
    table_names=g_table_columns_machine    
    for l in col_lst:
        del(table_names[l])

    if(len(table_names)>0):    
        for k,v in table_names.items():
            insert_table_column(db_name, table_name, k)
            create_table_ivalue_machine(db_name, k, v)
    
# create table
def create_table_ivalue_node(db_name, table_name, values):
    if((len(table_name)>0) & (len(values)>0)):
        # db connect
        conn = connect(db_name)
        
        try:
            with conn.cursor() as cursor:
                sql='CREATE TABLE IF NOT EXISTS ivalue_node_'+ table_name + ''' (
                    	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                    	`id_node` char(6) NOT NULL DEFAULT '' COMMENT 'node 번호' COLLATE 'utf8_general_ci',                	
                    	`total` FLOAT NOT NULL COMMENT '종합 건강도','''
                        
                for v in values:
                    sql += v + ' FLOAT NOT NULL,'
                        
                sql += ''' INDEX `id_node` (`id_node`) USING BTREE,
                    	INDEX `analysis_time` (`analysis_time`) USING BTREE
                        )
                        COMMENT='node 기준 진단 그룹 상세 결과 테이블'
                        COLLATE='utf8_general_ci'
                        ENGINE=InnoDB
                       '''
                cursor.execute(sql)        
            conn.commit()
        except Exception as e:
            logger_dbms.exception(e)
        finally:        
            conn.close()
    else:
        logger_dbms.warning('invalid para. length: table_name(%d), values(%d)', len(table_name), len(values))


def set_ivalue_node_table(db_name):
    table_names=g_table_columns_node

    for k,v in table_names.items():
        create_table_ivalue_node(db_name, k, v)

    
def update_node_table(db_name):
    '''
    Notice: . g_table_columns_node 에 추가할 항목을 우선 업데이트 후 실행 필요
            . 추가 업데이트만 가능
              g_# list에서 현재 db 항목 삭제 후 남은(추가 할) 항목 table 생성
    '''
    table_name='index_node'
    col_lst=get_table_columns(db_name, table_name).columns.to_list()
    del(col_lst[0:3]) # analysis_time, id_node, total
    
    table_names=g_table_columns_node    
    for l in col_lst:
        del(table_names[l])

    if(len(table_names)>0):    
        for k,v in table_names.items():
            insert_table_column(db_name, table_name, k)
            create_table_ivalue_node(db_name, k, v)
    
    
# create table
def create_table_ivalue_station(db_name, table_name, values):
    if((len(table_name)>0) & (len(values)>0)):
        # db connect
        conn = connect(db_name)
        
        try:
            with conn.cursor() as cursor:
                sql='CREATE TABLE IF NOT EXISTS ivalue_station_'+ table_name + ''' (
                    	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                    	`id_station` char(6) NOT NULL DEFAULT '' COMMENT 'Station 번호' COLLATE 'utf8_general_ci',                	
                    	`total` FLOAT NOT NULL COMMENT '종합 건강도','''
                        
                for v in values:
                    sql += v + ' FLOAT NOT NULL,'
                    logger_dbms.info(f'Create table: {table_name} -> {v}')
                        
                sql += ''' INDEX `id_station` (`id_station`) USING BTREE,
                    	INDEX `analysis_time` (`analysis_time`) USING BTREE
                        )
                        COMMENT='station 기준 진단 그룹 상세 결과 테이블'
                        COLLATE='utf8_general_ci'
                        ENGINE=InnoDB
                       '''
                cursor.execute(sql)
            conn.commit()
        except Exception as e:
            logger_dbms.exception(e)
        finally:        
            conn.close()
    else:
        logger_dbms.warning('invalid para. length: table_name(%d), values(%d)', len(table_name), len(values))



def set_ivalue_station_table(db_name):
    table_names=g_table_columns_station

    for k,v in table_names.items():
        create_table_ivalue_station(db_name, k, v)


def update_station_table(db_name):
    '''
    Notice: . g_table_columns_station 에 추가할 항목을 우선 업데이트 후 실행 필요
            . 추가 업데이트만 가능
              g_# list에서 현재 db 항목 삭제 후 남은(추가 할) 항목 table 생성
    '''
    table_name='index_station'
    col_lst=get_table_columns(db_name, table_name).columns.to_list()
    del(col_lst[0:3]) # analysis_time, id_node, total
    
    table_names=g_table_columns_station
    for l in col_lst:
        del(table_names[l])

    if(len(table_names)>0):    
        for k,v in table_names.items():
            insert_table_column(db_name, table_name, k)
            create_table_ivalue_station(db_name, k, v)


def create_table_cpk_ref_data(db_name):
# db connect
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `cpk_ref_data` (
                	`key_name` VARCHAR(64) NOT NULL COMMENT 'data key name',
                    `event_type` VARCHAR(16) NOT NULL COMMENT 'DRV or TRANS or STS',
                    `event_id` VARCHAR(64) NOT NULL COMMENT 'Event ID',
                    `start_end` VARCHAR(6) NOT NULL COMMENT 'Start or End or -',
                    `value_no` TINYINT UNSIGNED NOT NULL COMMENT 'Key value no 1~',
                    `mean` FLOAT NOT NULL COMMENT '',
                    `std` FLOAT NOT NULL COMMENT '',
                    `min` FLOAT NOT NULL COMMENT '',
                    `max` FLOAT NOT NULL COMMENT '',
                    `LSL` FLOAT NOT NULL COMMENT 'Lower Specification Limit',
                    `USL` FLOAT NOT NULL COMMENT 'Upper Specification Limit',                    
                    
                    INDEX `key_name` (`key_name`) USING BTREE
                )
                COMMENT='reference data table'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()


def create_table_cpk_machine(db_name):
    conn = connect(db_name)
    
    try:
        with conn.cursor() as cursor:
            sql='''
                CREATE TABLE IF NOT EXISTS `cpk_machine` (
                	`analysis_time` char(19) NOT NULL COMMENT '분석 일자시간',
                	`id_machine` char(6) NOT NULL DEFAULT '' COMMENT 'OHT Name' COLLATE 'utf8_general_ci',
                	`key_name` VARCHAR(64) NOT NULL COMMENT 'data key name',
                    `count` FLOAT NOT NULL COMMENT '',
                    `mean` FLOAT NOT NULL COMMENT '',
                    `std` FLOAT NOT NULL COMMENT '',
                    `min` FLOAT NOT NULL COMMENT '',
                    `max` FLOAT NOT NULL COMMENT '',
                    `K` FLOAT NOT NULL COMMENT '',
                    `Cp` FLOAT NOT NULL COMMENT '',
                    `Cpk` FLOAT NOT NULL COMMENT '',
                    `LSL` FLOAT NOT NULL COMMENT 'Lower Specification Limit',
                    `USL` FLOAT NOT NULL COMMENT 'Upper Specification Limit',

                	INDEX `analysis_time` (`analysis_time`) USING BTREE
                    INDEX `id_machine` (`id_machine`) USING BTREE
                )
                COMMENT='cpk result table by machine'
                COLLATE='utf8_general_ci'
                ENGINE=InnoDB
            '''
            cursor.execute(sql)        
        conn.commit()
    except Exception as e:
        logger_dbms.exception(e)
    finally:        
        conn.close()

def create_table(db_name):
    create_table_history(db_name)
    create_table_index_machine(db_name)
    create_table_index_node(db_name)
    create_table_index_station(db_name)
    set_ivalue_machine_table(db_name)
    set_ivalue_node_table(db_name)
    set_ivalue_station_table(db_name)
    
    create_table_cpk_ref_data(db_name)
    create_table_cpk_machine(db_name)


def update_table(db_name):
    update_machine_table(db_name)
    update_node_table(db_name)
    update_station_table(db_name)    

## 2. Insert Row(Data) #################################################################
def insert_history(db_name, df):
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
    df.to_sql(name='history', con=engine, if_exists='append', index=False)    
    
def insert_index_machine(db_name, df):
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
    df.to_sql(name='index_machine', con=engine, if_exists='append', index=False)    
    
def insert_df2table(db_name, df, table_name):
    logger_dbms.info('Insert df to table: %s', table_name)
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
    df.to_sql(name=table_name, con=engine, if_exists='append', index=False)    
    
def insert_df2grouptable(db_name, df, group_name, table_name, rename_columns, str_analysis_time):
    logger_dbms.info('Start update table of db: %s', table_name)
    
    df_db=df.rename(columns=rename_columns)
    df_db['analysis_time']=str_analysis_time
    
    if(group_name=='machine'):
        df_db=df_db.assign(id_machine=df_db.index.to_list())
    elif(group_name=='node'):
        df_db=df_db.assign(id_node=df_db.index.to_list())
    elif(group_name=='station'):
        df_db=df_db.assign(id_station=df_db.index.to_list())
    else:
        logger_dbms.warning('(insert_df2grouptable)Invalid Group name: ', group_name)
        return
    
    df_db=df_db.fillna(1000)
    
    df=get_table_columns(db_name, table_name)
    if(len(df.columns) == len(df_db.columns)):
        insert_df2table(db_name, df_db, table_name)
        logger_dbms.info('Success..!!')
    else:
        logger_dbms.warning('(%s) table missmatch columns count: Set/Real(%d / %d)',\
                            table_name, len(df_db.columns), len(df.columns))

## 3. Update #################################################################
def update_df2table(db_name, df, table_name):
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
    df.to_sql(name=table_name, con=engine, if_exists='replace', index=False)    

#전체 Select 하여 엑셀파일 쓰기
def select_all_to_excel(db_name, filepath, filename):
    conn = connect(db_name)
    # table list
    sql='SHOW TABLES' 
    df=pd.read_sql_query(sql, conn)
    lst_table=tuple(df['Tables_in_mcc'].tolist())

    wb = Workbook()
    ws = wb.active
    ws['A1']='mcc db tables to excel exported'
    try:
        with conn.cursor() as curs:
            for t in lst_table:
                logger_dbms.info('%s', t)
                # sheet title
                ws = wb.create_sheet()
                if(len(t)>30):
                    ws.title = t[:30]
                else:
                    ws.title = t
                
                # first row - fields name                
                sql='SHOW COLUMNS from '+ t
                logger_dbms.debug('SQL: %s', sql)
                df=pd.read_sql_query(sql, conn)
                cols=tuple(df['Field'].tolist())
                ws.append(cols)
                
                # export table                
                sql = "select * from " + t
                curs.execute(sql)
                rs = curs.fetchall()
                for row in rs:
                    ws.append(row)
                time.sleep(0)
    finally:
        conn.close()
        wb.close()
        
    wb.save(filepath+filename+'.xlsx')


def set_init_cpk_ref_data(db_name):
    df=get_table_columns(db_name, 'cpk_ref_data')
    
    lst_items=[{'key_name':'hand_release_movetime','event_type':'TRANS',\
                'event_id':'LD RELEASE','start_end':'End','value_no':2},\
               {'key_name':'hand_grip_movetime','event_type':'TRANS',\
                'event_id':'ULD GRIP','start_end':'End','value_no':2},\
               {'key_name':'shutter_front_open_movetime','event_type':'DRV',\
                'event_id':'SHUTTER FRONT OPEN','start_end':'End','value_no':2},\
               {'key_name':'shutter_front_close_movetime','event_type':'DRV',\
                'event_id':'SHUTTER FRONT CLOSE','start_end':'End','value_no':2},\
               {'key_name':'shutter_rear_open_movetime','event_type':'DRV',\
                'event_id':'SHUTTER REAR OPEN','start_end':'End','value_no':2},\
               {'key_name':'shutter_rear_close_movetime','event_type':'DRV',\
                'event_id':'SHUTTER REAR CLOSE','start_end':'End','value_no':2}]
    
    df=df.append(lst_items)
    df['mean']=df['std']=df['min']=df['max']=df['LSL']=df['USL']=0.0
    
    update_df2table(db_name, df, 'cpk_ref_data')

    
def get_cpk_ref_data(db_name):
    db_url="mysql://root:semes_12@localhost/"+db_name
    engine=sqlalchemy.create_engine(db_url,encoding='utf8')
    df=pd.read_sql_table('cpk_ref_data', con=engine)
    
    return df
    


#############################################################################################################
if __name__ == '__main__': 
    # Test
    #connect()
    #connect(db_name)
    
    # db create.
    db_name='mcc'
    print('Create db *' + db_name + '* start..!!')
    create_db(db_name)
    create_table(db_name)
    print('Create complete..!!')

    # db update
    update_table(db_name)
